from __future__ import annotations

from typing import Optional
from pathlib import Path
import pandas as pd

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.drawing.image import Image as XLImage

def export_xlsx(df: pd.DataFrame, out_path: str, logo_path: Optional[str] = None) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório"

    # Logo (opcional)
    if logo_path and Path(logo_path).exists():
        try:
            img = XLImage(logo_path)
            img.height = 60
            img.width = 180
            ws.add_image(img, "A1")
        except Exception:
            pass

    start_row = 5

    # Header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5597")
    for j, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=start_row, column=j, value=str(col))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Rows
    for i, row in enumerate(df.itertuples(index=False), start=start_row+1):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)

    # Auto width
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 45)

    wb.save(out_path)
